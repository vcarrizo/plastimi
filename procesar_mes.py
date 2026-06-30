#!/usr/bin/env python3
"""
procesar_mes.py
===============
Procesa los archivos rpcp199 (consumo de componentes por OP) y ragr004
(costos agrupados por cuenta contable) y genera:
  - datos/YYYY-MM.json    → datos del mes para el dashboard
  - datos/manifest.json   → índice de meses disponibles

Uso:
  python procesar_mes.py --rpcp199 <archivo.xlsx> --ragr004 <archivo.xlsx> --periodo 2026-05
  python procesar_mes.py --rpcp199 rpcp199.xlsx --ragr004 ragr004.xlsx --periodo 2026-06
"""

import argparse
import json
import os
import sys
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: Instalá openpyxl con: pip install openpyxl")
    sys.exit(1)

# ── TIPOS DE ARTÍCULO INCLUIDOS ──────────────────────────────────────────────
TIPOS_INCLUIDOS = {'2 - SIN IMPRESION', '3 - BOBINAS EXTRUSADAS'}

MESES_ES = {
    '01':'Enero','02':'Febrero','03':'Marzo','04':'Abril',
    '05':'Mayo','06':'Junio','07':'Julio','08':'Agosto',
    '09':'Septiembre','10':'Octubre','11':'Noviembre','12':'Diciembre'
}

# ── 1. PARSEAR ragr004 ────────────────────────────────────────────────────────
def parsear_ragr004(path):
    print(f"  Leyendo ragr004: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    unit_costs = {}   # "163.1" -> ARS/KG
    sections = {
        '02.01.01': 'Grumo',
        '02.01.02': 'Bobinas Tercerizadas',
        '02.01.03': 'Tintas Base',
        '02.03.03': 'Insumos',
    }
    current_section = None

    for row in ws.iter_rows(values_only=True):
        b, c = row[1], row[2]
        cant, valor = row[7], row[13]

        if b and isinstance(b, str):
            for prefix in sections:
                if prefix in b:
                    current_section = prefix
                    break

        if current_section and c and isinstance(c, str):
            code_part = c.split(' - ')[0].strip()
            prefix = current_section + '.'
            if code_part.startswith(prefix):
                item_code = code_part[len(prefix):]
                if cant and cant != 0 and valor:
                    unit_costs[item_code] = round(valor / cant, 6)
                # tintas: cant=0 pero tienen valor → costo unitario no calculable

    print(f"    → {len(unit_costs)} costos unitarios cargados")
    return unit_costs


# ── 2. PARSEAR Catálogo Iniflex ───────────────────────────────────────────────
def parsear_catalogo(path):
    print(f"  Leyendo catálogo: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    # Buscar hoja con datos de artículos
    ws = None
    for name in wb.sheetnames:
        sheet = wb[name]
        if sheet.max_row > 100:
            ws = sheet
            break
    if not ws:
        ws = wb[wb.sheetnames[0]]

    catalog = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        code = str(int(row[0])) if isinstance(row[0], float) else str(row[0])
        catalog[code] = {
            'tipo':     str(row[2] or ''),
            'grupo':    str(row[3] or ''),
            'subgrupo': str(row[4] or ''),
            'desc':     str(row[1] or ''),
        }

    tipo_count = defaultdict(int)
    for v in catalog.values():
        tipo_count[v['tipo']] += 1
    print(f"    → {len(catalog)} artículos | tipos: {dict(tipo_count)}")
    return catalog


# ── 3. PARSEAR rpcp199 ────────────────────────────────────────────────────────
def parsear_rpcp199(path, catalog, unit_costs):
    print(f"  Leyendo rpcp199: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    records = []
    current_op = None
    current_op_princ = None
    current_article = None
    etapas = set()

    for row in ws.iter_rows(values_only=True):
        b = row[1]

        if b == 'Anotaciónes:':
            etapa = row[25] if len(row) > 25 else None
            if etapa:
                etapas.add(str(etapa))
            continue

        if b == 'OP:':
            current_op = int(row[4]) if row[4] else None
            current_op_princ = int(row[10]) if row[10] else None
            current_article = str(row[19]) if row[19] else None
            continue

        if not (hasattr(b, 'year') and current_article):
            continue

        componente = row[16]
        cantidad = row[23]
        if not componente or not cantidad:
            continue

        # Parse artículo
        art_parts = current_article.split(' - ', 1)
        art_code_ver = art_parts[0].strip()
        art_code = art_code_ver.split('/')[0]
        art_name = art_parts[1].strip() if len(art_parts) > 1 else ''

        # Filtrar por tipo
        cat = catalog.get(art_code, {})
        tipo = cat.get('tipo', '')
        if tipo not in TIPOS_INCLUIDOS:
            continue

        # Parse componente
        comp_parts = str(componente).split(' - ', 1)
        comp_code_ver = comp_parts[0].strip()
        comp_code = comp_code_ver.replace('/', '.')
        comp_name = comp_parts[1].split(' / ')[0].strip() if len(comp_parts) > 1 else comp_code_ver

        # Costo
        costo_unit = unit_costs.get(comp_code)
        costo_total = round(costo_unit * cantidad, 2) if costo_unit else 0.0

        records.append({
            'fecha':        b.strftime('%Y-%m-%d'),
            'op':           current_op,
            'op_princ':     current_op_princ,
            'art_code':     art_code,
            'art_code_ver': art_code_ver,
            'art_name':     art_name,
            'art_tipo':     tipo,
            'art_grupo':    cat.get('grupo', ''),
            'comp_code':    comp_code,
            'comp_name':    comp_name,
            'cantidad_kg':  round(cantidad, 4),
            'costo_unit':   round(costo_unit, 4) if costo_unit else None,
            'costo_total':  costo_total,
            'sin_costo':    costo_unit is None,
        })

    print(f"    → {len(records)} registros | etapas: {etapas}")
    return records


# ── 4. AGREGAR Y CONSTRUIR JSON ───────────────────────────────────────────────
def construir_dashboard_data(records, periodo):
    by_art = defaultdict(lambda: {
        'kg': 0, 'costo': 0, 'ops': set(),
        'comps': defaultdict(lambda: {'kg': 0, 'costo': 0})
    })
    by_comp = defaultdict(lambda: {'kg': 0, 'costo': 0})
    by_grupo = defaultdict(lambda: {'kg': 0, 'costo': 0, 'arts': set()})

    for r in records:
        key = r['art_code_ver']
        by_art[key]['kg']        += r['cantidad_kg']
        by_art[key]['costo']     += r['costo_total']
        by_art[key]['ops'].add(r['op'])
        by_art[key]['comps'][r['comp_name']]['kg']    += r['cantidad_kg']
        by_art[key]['comps'][r['comp_name']]['costo'] += r['costo_total']
        by_art[key]['name']  = r['art_name']
        by_art[key]['tipo']  = r['art_tipo']
        by_art[key]['grupo'] = r['art_grupo']

        by_comp[r['comp_name']]['kg']    += r['cantidad_kg']
        by_comp[r['comp_name']]['costo'] += r['costo_total']

        g = r['art_grupo'].split(' - ', 1)[-1] if ' - ' in r['art_grupo'] else r['art_grupo']
        by_grupo[g]['kg']    += r['cantidad_kg']
        by_grupo[g]['costo'] += r['costo_total']
        by_grupo[g]['arts'].add(r['art_code_ver'])

    top_arts  = sorted(by_art.items(),  key=lambda x: -x[1]['costo'])[:15]
    top_comps = sorted(by_comp.items(), key=lambda x: -x[1]['costo'])[:15]
    grupos_sorted = sorted(by_grupo.items(), key=lambda x: -x[1]['costo'])

    total_kg    = sum(r['cantidad_kg']  for r in records)
    total_costo = sum(r['costo_total']  for r in records)
    total_ops   = len(set(r['op']       for r in records))
    total_arts  = len(by_art)

    anio, mes = periodo.split('-')
    label_periodo = f"{MESES_ES.get(mes, mes)} {anio}"

    summary = {
        'periodo':      periodo,
        'label':        label_periodo,
        'etapa':        'Extrusión Parque',
        'total_kg':     round(total_kg, 0),
        'total_costo':  round(total_costo, 0),
        'total_ops':    total_ops,
        'total_arts':   total_arts,
        'costo_por_kg': round(total_costo / total_kg, 2) if total_kg else 0,
    }

    # Stacked bar: top15 arts × componentes
    top15_codes = [k for k, _ in top_arts]
    all_comps_in_top = set()
    for code in top15_codes:
        for comp in by_art[code]['comps']:
            all_comps_in_top.add(comp)
    comp_order = sorted(all_comps_in_top, key=lambda c: -by_comp[c]['costo'])

    stacked_bar = {
        'art_labels':  [f"{k} · {by_art[k]['name'][:35]}" for k in top15_codes],
        'art_codes':   top15_codes,
        'comp_labels': comp_order,
        'datasets': [
            {'comp': comp, 'values': [
                round(by_art[code]['comps'].get(comp, {}).get('costo', 0), 0)
                for code in top15_codes
            ]}
            for comp in comp_order
        ],
    }

    # Per-article component detail
    art_comp_detail = {}
    for k, v in by_art.items():
        comps_sorted = sorted(v['comps'].items(), key=lambda x: -x[1]['costo'])
        art_comp_detail[k] = {
            'name':  v['name'],
            'grupo': v['grupo'].split(' - ', 1)[-1] if ' - ' in v['grupo'] else v['grupo'],
            'kg':    round(v['kg'], 0),
            'costo': round(v['costo'], 0),
            'ops':   len(v['ops']),
            'comps': [
                {
                    'name':  c,
                    'kg':    round(d['kg'], 2),
                    'costo': round(d['costo'], 0),
                    'pct':   round(d['costo'] / v['costo'] * 100, 1) if v['costo'] else 0,
                }
                for c, d in comps_sorted if d['costo'] > 0
            ],
        }

    data = {
        'summary': summary,
        'by_grupo': [
            {'grupo': k, 'kg': round(v['kg'], 0), 'costo': round(v['costo'], 0), 'arts': len(v['arts'])}
            for k, v in grupos_sorted
        ],
        'top_arts': [
            {
                'code':     k,
                'name':     v['name'][:55],
                'tipo':     v['tipo'],
                'grupo':    v['grupo'].split(' - ', 1)[-1] if ' - ' in v['grupo'] else v['grupo'],
                'kg':       round(v['kg'], 0),
                'costo':    round(v['costo'], 0),
                'ops':      len(v['ops']),
                'costo_kg': round(v['costo'] / v['kg'], 2) if v['kg'] else 0,
            }
            for k, v in top_arts
        ],
        'top_comps': [
            {
                'name':     k[:50],
                'kg':       round(v['kg'], 0),
                'costo':    round(v['costo'], 0),
                'costo_kg': round(v['costo'] / v['kg'], 2) if v['kg'] else 0,
            }
            for k, v in top_comps
        ],
        'stacked_bar':     stacked_bar,
        'art_comp_detail': art_comp_detail,
        'detail': [
            {
                'fecha':      r['fecha'],
                'op':         r['op'],
                'art_code':   r['art_code_ver'],
                'art_name':   r['art_name'][:55],
                'art_grupo':  r['art_grupo'].split(' - ', 1)[-1] if ' - ' in r['art_grupo'] else r['art_grupo'],
                'comp_name':  r['comp_name'][:40],
                'kg':         r['cantidad_kg'],
                'costo_unit': r['costo_unit'],
                'costo_total':r['costo_total'],
                'sin_costo':  r['sin_costo'],
            }
            for r in records
        ],
    }

    return data


# ── 5. ACTUALIZAR MANIFEST ────────────────────────────────────────────────────
def actualizar_manifest(output_dir, periodo, summary):
    manifest_path = os.path.join(output_dir, 'manifest.json')
    manifest = {}
    if os.path.exists(manifest_path):
        with open(manifest_path, 'r', encoding='utf-8') as f:
            manifest = json.load(f)

    manifest[periodo] = {
        'periodo':      periodo,
        'label':        summary['label'],
        'total_kg':     summary['total_kg'],
        'total_costo':  summary['total_costo'],
        'total_arts':   summary['total_arts'],
        'total_ops':    summary['total_ops'],
        'procesado':    datetime.now().strftime('%Y-%m-%d %H:%M'),
        'archivo':      f"{periodo}.json",
    }

    # Ordenar por período descendente
    manifest_sorted = dict(sorted(manifest.items(), reverse=True))

    with open(manifest_path, 'w', encoding='utf-8') as f:
        json.dump(manifest_sorted, f, ensure_ascii=False, indent=2)

    print(f"    → manifest.json actualizado ({len(manifest_sorted)} períodos)")
    return manifest_sorted


# ── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description='Procesador mensual de costos Plastimi')
    parser.add_argument('--rpcp199',  required=True, help='Archivo rpcp199 (.xlsx)')
    parser.add_argument('--ragr004',  required=True, help='Archivo ragr004 (.xlsx)')
    parser.add_argument('--catalogo', default=None,  help='Catálogo de productos Iniflex (.xlsx)')
    parser.add_argument('--periodo',  required=True, help='Período en formato YYYY-MM (ej: 2026-05)')
    parser.add_argument('--output',   default='datos', help='Carpeta de salida (default: datos/)')
    args = parser.parse_args()

    # Validar período
    try:
        anio, mes = args.periodo.split('-')
        assert len(anio) == 4 and len(mes) == 2
        int(anio); int(mes)
    except Exception:
        print("ERROR: --periodo debe ser YYYY-MM (ej: 2026-05)")
        sys.exit(1)

    # Validar archivos
    for f, name in [(args.rpcp199, 'rpcp199'), (args.ragr004, 'ragr004')]:
        if not os.path.exists(f):
            print(f"ERROR: No se encontró el archivo {name}: {f}")
            sys.exit(1)

    # Catálogo: si no se especifica, buscar en la misma carpeta
    catalogo_path = args.catalogo
    if not catalogo_path:
        for candidate in ['Productos Iniflex.xlsx', 'catalogo.xlsx', 'productos.xlsx']:
            if os.path.exists(candidate):
                catalogo_path = candidate
                break
    if not catalogo_path or not os.path.exists(catalogo_path):
        print("ERROR: No se encontró el catálogo de productos.")
        print("  Especificalo con --catalogo <archivo.xlsx>")
        print("  O colocá 'Productos Iniflex.xlsx' en la misma carpeta.")
        sys.exit(1)

    print(f"\n{'='*55}")
    print(f"  Plastimi · Procesamiento período {args.periodo}")
    print(f"{'='*55}")

    # Crear carpeta de salida
    os.makedirs(args.output, exist_ok=True)

    # Procesar
    print("\n[1/4] Cargando costos unitarios (ragr004)...")
    unit_costs = parsear_ragr004(args.ragr004)

    print("\n[2/4] Cargando catálogo de productos...")
    catalog = parsear_catalogo(catalogo_path)

    print("\n[3/4] Procesando consumo de componentes (rpcp199)...")
    records = parsear_rpcp199(args.rpcp199, catalog, unit_costs)

    print("\n[4/4] Construyendo JSON del dashboard...")
    data = construir_dashboard_data(records, args.periodo)

    # Guardar JSON del período
    json_path = os.path.join(args.output, f"{args.periodo}.json")
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, separators=(',', ':'))

    json_kb = os.path.getsize(json_path) / 1024
    print(f"    → {json_path} ({json_kb:.0f} KB)")

    # Actualizar manifest
    manifest = actualizar_manifest(args.output, args.periodo, data['summary'])

    print(f"\n{'='*55}")
    print(f"  ✓ Período {args.periodo} procesado exitosamente")
    print(f"  ✓ {len(records)} registros | {data['summary']['total_arts']} artículos")
    print(f"  ✓ {data['summary']['total_kg']:,.0f} KG | ARS {data['summary']['total_costo']:,.0f}")
    print(f"\n  Archivos generados:")
    print(f"    → {json_path}")
    print(f"    → {os.path.join(args.output, 'manifest.json')}")
    print(f"\n  Subir al repo:")
    print(f"    git add {args.output}/")
    print(f"    git commit -m 'datos: agregar período {args.periodo}'")
    print(f"    git push")
    print(f"{'='*55}\n")


if __name__ == '__main__':
    main()
